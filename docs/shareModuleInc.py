from datetime import datetime, timedelta
from django.db import connection
from .models import *
from openpyxl import load_workbook
import tabula
import os


def getincplanbyclassif(organization, date = None, classification = ''):
    usl = ""
    if not (classification == None or classification == ''):
        usl = usl + " and _classification_id=" + str(classification)

    if date == None:
        date = datetime.now()

    date_start = datetime.strptime(str(date.year) + "-01-01 00:00:00", '%Y-%m-%d %H:%M:%S')
    dateend = date

    query = f"""with utv as (SELECT _classification_id, sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12 
                                FROM public.docs_utv_inc_tbl1
                                where _organization_id = {organization} and not deleted {usl} and _date>='{date_start}' and _date <= '{dateend}' 
                                group by _classification_id),
                        izm as (SELECT _classification_id,  sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12 
                                FROM public.docs_izm_inc_tbl1
                                where _organization_id = {organization} and not deleted {usl} and _date>='{date_start}' and _date <= '{dateend}'
                                group by _classification_id),
                        union_utv_izm as (select * from utv
                                            union all
                                            select * from izm),
                        classname as (SELECT * FROM public.dirs_classification_income
                                     WHERE id in (select _classification_id from union_utv_izm))
                    SELECT _classification_id, CONCAT(classname.code, ' - ', classname.name_rus) as classification_name,  sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12 from union_utv_izm
                    LEFT JOIN classname
                    ON _classification_id = classname.id
                    GROUP BY _classification_id, classification_name
                    ORDER BY _classification_id"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                for row in cursor.fetchall()]
    return result


def pdftotext():
    listnum = ('0', '1', '2', '3', '4', '5', '6', '7','8', '9')
    # Извлечение таблицы из PDF
    df = tabula.read_pdf('pdf219.pdf', pages='all')

    kp_code = ''
    index = 0
    budjet_code = ''
    itog2 = 0

    for page in df:
        page.to_excel('330410.xlsx', index=False)
   
        # Загружаем книгу Excel
        workbook = load_workbook('330410.xlsx')
        # Получаем список названий листов в книге
        sheet_names = workbook.sheetnames
        # Выбираем первый лист
        first_sheet = workbook[sheet_names[0]]

        # Читаем данные из ячеек в выбранном листе
        rowcount = 0
        for row in first_sheet.iter_rows(values_only=True):
            str_0 = row[0]
            str_1 = row[1]
            

            if not str_0 == None and len(str_0)>=6:
                mas_0 = str_0[0:6]
                if mas_0[0] in listnum and  mas_0[5] in listnum:
                    kp_code = mas_0
                    # print(kp_code)
                    continue

            
            if str_1 == None and rowcount == 1:
                kp_code = ''
       
            if not kp_code == '':
                if str_1 == None or len(str_1)<6:
                    continue            
                mas_1 = str_1[0:6]
                if mas_1[0] in listnum and mas_1[5] in listnum and mas_1[1] in listnum and mas_1[2] in listnum and mas_1[3] in listnum and mas_1[4] in listnum:
                    rowcount = 1
                    index = index + 1
                    budjet_code = mas_1
                    print(str(index) +') ' + kp_code + ' ' + budjet_code + ' ' + str(row[2])+ ' ' + str(row[3])+ ' ' + str(row[4])+ ' ' + str(row[5])+ ' ' + str(row[6])+ ' ' + str(row[7]))
                    # kp_code = ''
                    try:
                        itog2 = itog2 + float(str(row[2]).replace(',', ''))
                    except:
                        print('error ' + row[2])
                        break


                    # if first_sheet._cells[rowcount + 1, 2].value == None:
                    #     kp_code = ''
                    #     budjet_code = ''
                    # continue

                

                


        # Закрываем книгу
        workbook.close()
    print(itog2)
    return True



def fkrreadxls(path='fkr.xls'):
    listnum = ('0', '1', '2', '3', '4', '5', '6', '7','8', '9')

    if os.path.exists(path):  
        # Загружаем книгу Excel
        workbook = load_workbook(path)
        # Получаем список названий листов в книге
        sheet_names = workbook.sheetnames
        # Выбираем первый лист
        first_sheet = workbook[sheet_names[0]]

        # Читаем данные из ячеек в выбранном листе
        rowcount = 0
        for row in first_sheet.iter_rows(values_only=True):
            str_0 = row[0]
            str_1 = row[1]
            
        # Закрываем книгу
        workbook.close()
    return True






