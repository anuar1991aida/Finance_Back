from datetime import datetime, timedelta
from django.db import connection
from .models import *
from openpyxl import load_workbook
import tabula
import os




def fkrreadxls(path='fkrnewkaz.xlsx'):
    listnum = ('0', '1', '2', '3', '4', '5', '6', '7','8', '9')
    
    if os.path.exists(path):  
        obj_fg = funcgroup.objects.all().values_list()
        obj_fpg = funcpodgroup.objects.all().values_list()
        obj_prog = program.objects.all().values_list()
        obj_podprog = podprogram.objects.all().values_list()
        obj_fkr = fkr.objects.all().values_list()
 
        workbook = load_workbook(path)
        sheet_names = workbook.sheetnames
        first_sheet = workbook[sheet_names[0]]


        fg_id = 0
        fg_code = ''
        fpg_id = 0
        fpg_code = ''
        abp_id = 0
        abp_code = ''
        pr_id = 0
        pr_code = ''
        pr_name = ''
        ppr_id = 0
        ppr_code = ''

        lang = ''
        if first_sheet._cells[1,1].value == 'Функциональная группа':
            lang = 'rus'

        if first_sheet._cells[1,1].value == 'Функционалдық топ':
            lang = 'kaz'

        if lang == '':
            return 'error excel'
       

        if lang == 'kaz':
            # Читаем данные из ячеек в выбранном листе
            for row in first_sheet.iter_rows(values_only=True):

                # Пропускаем заголовки таблицы
                if type(row[0]) == str or type(row[1]) == str or type(row[2]) == str or type(row[3]) == str or type(row[4]) == str:
                    continue

                # 1. Находим функциональную группу
                if not row[0] == None:
                    # Если длина кода равна 1, то дополним впереди 0: типа 01, 02 и т.д.
                    if len(str(row[0])) == 1:
                        code_fg = '0' + str(row[0])
                    else:
                        code_fg = str(row[0])
                    
                    zapis = funcgroup.objects.get(code = code_fg)
                    zapis.name_kaz = row[5]
                    zapis.save()
                    fg_id = zapis.id
                    continue  #Переходим к следующему циклу
                    #----------------------------------------------------------------- 


                # 2. Находим функциональную Подгруппу
                if not row[1] == None:
                    zapis = funcpodgroup.objects.get(code = row[1], _funcgroup = fg_id)
                    zapis.name_kaz = row[5]
                    zapis.save()
                    fpg_id = zapis.id
                    code_fpg = str(row[1])
                    continue  #Переходим к следующему циклу

                
                # Находим АБП
                if not row[2] == None:
                    try:
                        zapis = abp.objects.get(code = row[2])
                        zapis.name_kaz = row[5]
                        zapis.save()
                        abp_id = zapis.id
                        code_abp = str(row[2])
                    except:
                        print('abp code:' + str(row[2]))
                    continue  #Переходим к следующему циклу
                    #-----------------------------------------------------------------
                 



                # Находим программы
                if not row[3] == None:
                    if len(str(row[3])) == 1:
                        code_pr = code_fg + '/' + code_fpg + '/' + code_abp + '/00' + str(row[3])
                    elif len(str(row[3])) == 2:
                        code_pr = code_fg + '/' + code_fpg + '/' + code_abp + '/0' + str(row[3])
                    else:
                        code_pr = code_fg + '/' + code_fpg + '/' + code_abp + '/' + str(row[3])

                    try:
                        zapis = program.objects.get(code = code_pr)
                        zapis.name_kaz = row[5]
                        zapis.save()
                        pr_id = zapis.id
                    except:
                        print('pr code:' + code_pr)
                    continue  #Переходим к следующему циклу
                    #-----------------------------------------------------------------



                # Находим Подпрограммы
                if not row[4] == None:
                    if len(str(row[4])) == 1:
                        code_ppr = code_pr + '/00' + str(row[4])
                    elif len(str(row[4])) == 2:
                        code_ppr = code_pr + '/0' + str(row[4])
                    else:
                        code_ppr = code_pr + '/' + str(row[4])

                    masscode = code_ppr.split('/')
                    codefkr = masscode[2] + '/' + masscode[3] + '/' + masscode[4]

                    try:
                        zapis = podprogram.objects.get(code = code_ppr)
                        zapis.name_kaz = row[5]
                        zapis.save()
                        ppr_id = zapis.id
                    except:
                        print('ppr code:' + code_ppr)

                    try:
                        zapis = fkr.objects.get(code = codefkr)
                        zapis.name_kaz = row[5]
                        zapis.save()
                    except:
                        print('ppr code:' + codefkr)
                    continue  #Переходим к следующему циклу


            

        if lang == 'rus':
             # Читаем данные из ячеек в выбранном листе
            for row in first_sheet.iter_rows(values_only=True):

                # Пропускаем заголовки таблицы
                if type(row[0]) == str or type(row[1]) == str or type(row[2]) == str or type(row[3]) == str or type(row[4]) == str:
                    continue

                # 1. Находим функциональную группу
                if not row[0] == None:
                    # Если длина кода равна 1, то дополним впереди 0: типа 01, 02 и т.д.
                    if len(str(row[0])) == 1:
                        code = '0' + str(row[0])
                    else:
                        code = str(row[0])
                    # -----------------------------------------------------------------

                    # Ищем, есть ли в базе функ группа с таким кодом
                    existfg = False
                    for itemfg in obj_fg:
                        if itemfg[1] == code:
                            existfg = True
                            fg_id = itemfg[0]
                            fg_code = itemfg[1]
                    # -----------------------------------------------

                    # Если не существует, то создаем новую функциональную группу
                    if not existfg:
                        funcgroupd = funcgroup()
                        funcgroupd.code = code
                        funcgroupd.name_rus = row[5]
                        funcgroupd.save()
                        fg_id = funcgroupd.id
                        fg_code = code
                    # ---------------------------------------------- 

                


                # 2. Находим функциональную Подгруппу
                if not row[1] == None:
                    code = str(row[1])
                    existfg = False
                    for itemfpg in obj_fpg:
                        if itemfpg[1] == code:
                            existfg = True
                            fpg_id = itemfpg[0]
                            fpg_code = itemfpg[1]
                    if not existfg:
                        funcpodgroupd = funcpodgroup()
                        funcpodgroupd.code = code
                        funcpodgroupd.name_rus = row[5]
                        funcpodgroupd._funcgroup_id = fg_id
                        funcpodgroupd.save()
                        fpg_id = funcpodgroupd.id
                        fpg_code = funcpodgroupd.code

                
                # Находим АБП
                if not row[2] == None:
                    code = str(row[2])
                    existfg = False
                    obj_abp = abp.objects.all().values_list()
                    for itemabp in obj_abp:
                        if itemabp[1] == code:
                            existfg = True
                            abp_id = itemabp[0]
                            abp_code = itemabp[1]
                    if not existfg:
                        abpd = abp()                      
                        abpd.code = code
                        abpd.name_rus = row[5]
                        abpd.save()
                        abp_id = abpd.id
                        abp_code = abpd.code



                # Находим программы
                if not row[3] == None:
                    if len(str(row[3])) == 1:
                        code = fg_code + '/' + fpg_code + '/' + abp_code + '/00' + str(row[3])
                    elif len(str(row[3])) == 2:
                        code = fg_code + '/' + fpg_code + '/' + abp_code + '/0' + str(row[3])
                    else:
                        code = fg_code + '/' + fpg_code + '/' + abp_code + '/' + str(row[3])

 
                    existfg = False
                    for itempr in obj_prog:
                        if (itempr[1]) == code:
                            existfg = True
                            pr_id = itempr[0]
                            pr_code = itempr[1]
                            pr_name = itempr[3]
                    if not existfg:
                        progd = program()
                        progd.code =  code
                        progd.name_rus = row[5]
                        progd._funcgroup = fg_id
                        progd._funcpodgroup_id = fpg_id
                        progd._abp_id = abp_id
                        progd.save()
                        pr_id = progd.id
                        pr_code = progd.code
                        pr_name = progd.name_rus



                # Находим Подпрограммы
                if not row[4] == None:
                    if len(str(row[4])) == 1:
                        code = pr_code + '/00' + str(row[4])
                    elif len(str(row[4])) == 2:
                        code = pr_code + '/0' + str(row[4])
                    else:
                        code = pr_code + '/' + str(row[4])


                    existfg = False
                    for itempodpr in obj_podprog:
                        if (itempodpr[1]) == code:
                            existfg = True
                            ppr_id = itempodpr[0]
                            ppr_code = itempodpr[1]

                    if not existfg:
                        podprogd = podprogram()
                        podprogd.code =  code
                        podprogd.name_rus = row[5]
                        podprogd._funcgroup = fg_id
                        podprogd._funcpodgroup = fpg_id
                        podprogd._abp = abp_id
                        podprogd._program_id = pr_id
                        podprogd.save()
                        ppr_id = podprogd.id
                        ppr_code = podprogd.code


                    # Создаем ФКР, имея уже все данные для создания
                    masscode = ppr_code.split('/')
                    codefkr = masscode[2] + '/' + masscode[3] + '/' + masscode[4]
                    existfg = False
                    for itemfkr in obj_fkr:
                        if (itemfkr[1]) == codefkr:
                            existfg = True

                    if not existfg:
                        fkrnew = fkr()
                        fkrnew.code =  codefkr
                        fkrnew.name_rus = pr_name + '(' + row[5] + ')'
                        fkrnew._funcgroup = fg_id
                        fkrnew._funcpodgroup = fpg_id
                        fkrnew._abp = abp_id
                        fkrnew._program_id = pr_id
                        fkrnew._podprogram_id = ppr_id
                        fkrnew.save()
                

       
        # Закрываем книгу
        workbook.close()
    return 'Успешно'






